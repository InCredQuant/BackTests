# -*- coding: utf-8 -*-
"""
Created on Fri Nov 10 10:29:54 2023
@author: Viren@Incred
This is for Long only Funds, where we are backtesting to take long positons on Cash Equity stocks based on some price based signals

indices = ['BSE100 INDEX', 'BSE200 INDEX', 'BSE500 INDEX', 'NSEMCAP INDEX', 'NSESMCP INDEX']
Selecting top 30 Stocks based on each factor.
1.	Weekly/Monthly  MACD of BSE200 Stocks
2.	Monthly Momentum : Last 1 Month Momentum, higehr is better
3.	Monthly Volatility
4.	Monthly Mean Reversion
5.	6 Months Momentum
6.	12-1 Month Momentum on BSE500, MidCap, Small Cap Index
7.	Stocks at 52 Week high
8.	Negatively/positively Skewed Stocks
9.	Persistence of Last 1 Yr ( WoW basis it should consistently Perform)
10.	1 Yr Momentum with Volatility Adjustment
11.	1 yr Mean Drawdown or Drawdown (lower is Better)
12.	3 Years price Volatility


"""

from FactoryBackTester import FactoryBackTester
#from FactoryBackTester import NameDecorator
import MyTechnicalLib
import pandas
import numpy
import datetime
import pdb
import matplotlib.pyplot as plt
import copy
from random import choice
#from scipy import stats
import pdb
import warnings
warnings.filterwarnings("ignore")

from order_base import Order, Position, OptionType, Segment
from trade_register import TradeRegister
from stats import Stats, Filter

class PriceMomentum(FactoryBackTester):
    def __init__(self,data, indexName, modelName, rebalType = 'm', stocksCount = [30, 50]):
        FactoryBackTester.__init__(self,data)
        self.IndexName = indexName
        self.ModelName = modelName
        self.RebalType = rebalType #'w', 'm', 'q', 'var',
        self.SelectionCount = stocksCount
        
    def basicdatainitialize(self):           
        #self.CurrentTime = pandas.to_datetime('2022-12-27')#min(list(self.BackTestData.IndexCompDict[self.IndexName].keys())) + datetime.timedelta(1)#pandas.to_datetime('2022-12-27')
        self.CurrentTime = min(list(self.BackTestData.IndexCompDict[self.IndexName].keys())) + datetime.timedelta(1)
        if self.RebalType.lower() == 'var':
            self.indexprice = self.BackTestData.indexprice.loc[self.CurrentTime:, self.IndexName].dropna()
            self.indexpriceSq = numpy.square(self.indexprice.pct_change(1)*1000)
            monthLastDates = [self.indexpriceSq.index[0]]
            cumsum = 0
            for it in self.indexpriceSq.iloc[1:].index:
                cumsum += self.indexpriceSq.loc[it]
                if cumsum/4000 >= 1:
                    monthLastDates.append(it)
                    cumsum = 0
        else:                    
            monthLastDates = self.BackTestData.Close.resample(self.RebalType.lower()).last().index
        self.UpdateDates = [max([jt for jt in  self.BackTestData.Close.index if jt <=it]) for it in monthLastDates if it >= self.CurrentTime]
        #self.UpdateDates.append('')
        #self.TransactionCostRate = 0.000#25# Total 2 bps Transaction Charges
        #self.StopLossLimit = -0.10
        #self.PositionExitDF = pandas.DataFrame(numpy.nan,index=self.Close.index,columns=self.Close.columns)
        #self.PositionWOSL = pandas.DataFrame(numpy.zeros_like(self.Close),index=self.Close.index,columns=self.Close.columns)
        self.order = Order()
        self.trade_reg = TradeRegister()
        
        self.DrawDown = self.Close/self.BackTestData.High.rolling(window = 252, min_periods = 126).max()
        self.DrawDownAvg = self.DrawDown.rolling(window = 252, min_periods = 126).mean()
        
        self.MACD, self.MACDSignal = MyTechnicalLib.MACD(self.Close.resample('w').last())
        self.MACDDiff = numpy.subtract(self.MACD, self.MACDSignal)
        
        #self.TurnOverAvg = self.BackTestData.TurnOver.rolling(window = 126, min_periods = 126).mean()
        #self.MCapAvg = self.BackTestData.MCap.rolling(window = 126, min_periods = 100).mean()
        #self.Skewness = self.Close.pct_change().rolling(252, min_periods = 126).skew()
        
        
    def declarecurrentvariables(self):
        self.LastPosition=self.Position.loc[self.LastTime]
        self.CurrentNAV=self.NAV.loc[self.CurrentTime,'NAV']
        self.CurrentPrice=self.Close.loc[self.CurrentTime].dropna()
        
        
    def detectupdatedate(self):
        if self.CurrentTime in self.UpdateDates:
            return True
    '''
    def StopLossHandler(self):
        self.PositionWOSL.loc[self.CurrentTime] = self.PositionWOSL.loc[self.LastTime]
        positionWOSL = self.PositionWOSL.loc[self.CurrentTime]
        tempDF = positionWOSL[positionWOSL != 0]
        
        for ticker in tempDF.index:
            #self.StopLoss(ticker, self.StopLossLimit)
            self.StopLossTrail(ticker, self.StopLossLimit)
        
        self.Position.loc[self.CurrentTime] = self.PositionWOSL.loc[self.CurrentTime]
        for ticker in tempDF.index:
            TradeTakenDate = self.DetectPostionStartDate(ticker, self.PositionWOSL)
            checkExitPosition = self.PositionExitDF.loc[TradeTakenDate:self.CurrentTime, ticker].dropna()
            if len(checkExitPosition) >0:
                oldPosNifty = self.Position.loc[self.CurrentTime, 'NIFTY INDEX']
                oldPosTicker = self.Position.loc[self.CurrentTime, ticker]
                newPos = oldPosNifty+ oldPosTicker
                
                self.Position.loc[self.CurrentTime, 'NIFTY INDEX'] = numpy.sign(newPos)
                if numpy.sign(oldPosNifty)*numpy.sign(oldPosTicker) == -1:
                    self.Position.loc[self.CurrentTime, 'Cash'] = 1
                    self.CapitalAllocation.loc[self.CurrentTime, 'Cash'] += (2*numpy.minimum(numpy.abs(self.CapitalAllocation.loc[self.CurrentTime, ticker]), numpy.abs(self.CapitalAllocation.loc[self.CurrentTime, 'NIFTY INDEX'])))
                self.CapitalAllocation.loc[self.CurrentTime, 'NIFTY INDEX'] += self.CapitalAllocation.loc[self.CurrentTime, ticker]
                self.CapitalAllocation.loc[self.CurrentTime, ticker] = 0
                self.Position.loc[self.CurrentTime, ticker] = 0
            else:
                self.Position.loc[self.CurrentTime, ticker] = self.PositionWOSL.loc[self.CurrentTime, ticker]
    '''
        
    def UpdateSpecificStats(self):
        namesinIndex = set.intersection( set(self.Close.columns), set(self.GetLatestIndexComponents()))
        self.StocksCount = len(namesinIndex)
        if self.IndexName == 'BSEALL INDEX':
            bse500Const = self.BackTestData.IndexCompDict['BSE500 INDEX'][max([ikey for ikey  in self.BackTestData.IndexCompDict['BSE500 INDEX'].keys() if ikey <= self.CurrentTime])]
            namesinIndex = set.difference( set(namesinIndex), set(bse500Const))
            self.StocksCount = len(namesinIndex)
        
        if self.ModelName == '1YrPriceMomVolAdj':
            PriceMom = self.Close.pct_change(252).loc[self.CurrentTime, namesinIndex].dropna()
            PriceVol = self.Close.pct_change().loc[self.CurrentTime - datetime.timedelta(365):self.CurrentTime, namesinIndex].std().dropna()
            #PriceVol = PriceVol.rank(ascending = False)/len(PriceVol)            
            PriceMomwithVol = PriceMom/PriceVol
            PriceMomwithVol = PriceMomwithVol.rank(ascending = True)/len(PriceMomwithVol)
            self.FactorRanks = PriceMomwithVol.dropna()
        elif self.ModelName == '6MPriceMom':
            PriceMom = self.Close.pct_change(126).loc[self.CurrentTime, namesinIndex].dropna()
            PriceMom = PriceMom.rank(ascending = True)/len(PriceMom)
            self.FactorRanks = PriceMom
        elif self.ModelName == '3MPriceMom':
            PriceMom = self.Close.pct_change(63).loc[self.CurrentTime, namesinIndex].dropna()
            PriceMom = PriceMom.rank(ascending = True)/len(PriceMom)
            self.FactorRanks = PriceMom
        elif self.ModelName == '1YrPriceMom':
            PriceMom = self.Close.pct_change(252).loc[self.CurrentTime, namesinIndex].dropna()
            PriceMom = PriceMom.rank(ascending = True)/len(PriceMom)
            self.FactorRanks = PriceMom
        elif self.ModelName == 'DrawDownAvgLow':
            drawdown = self.DrawDownAvg.loc[self.CurrentTime, namesinIndex].dropna()
            drawdown = drawdown.rank(ascending = True)/len(drawdown)
            self.FactorRanks = drawdown
        elif self.ModelName == '1YrPriceLowVol':
            PriceVol = self.Close.pct_change().loc[self.CurrentTime - datetime.timedelta(365):self.CurrentTime, namesinIndex].std().dropna()
            PriceVol = PriceVol.rank(ascending = False)/len(PriceVol)
            self.FactorRanks = PriceVol
        elif self.ModelName == '6MPriceLowVol':
            PriceVol = self.Close.pct_change().loc[self.CurrentTime - datetime.timedelta(183):self.CurrentTime, namesinIndex].std().dropna()
            PriceVol = PriceVol.rank(ascending = False)/len(PriceVol)
            self.FactorRanks = PriceVol
        elif self.ModelName == 'MonthlyMACD':
            macd = self.MACDDiff.loc[:self.CurrentTime + datetime.timedelta(4), namesinIndex].iloc[-1].dropna()
            macd = macd[macd>0].dropna()
            macd = macd.rank(ascending = True)/len(macd)
            self.FactorRanks = macd
        elif self.ModelName == 'Near52Wkhigh_NrTop15Pct':
            near52WkHigh = self.Close.loc[self.CurrentTime, namesinIndex]/self.BackTestData.High.loc[self.CurrentTime - datetime.timedelta(365) : self.CurrentTime, namesinIndex].max(axis = 0)
            near52WkHigh = near52WkHigh[near52WkHigh>=0.85].dropna()
            near52WkHigh = near52WkHigh.rank(ascending = True)/len(near52WkHigh)
            self.FactorRanks = near52WkHigh
        
    def updateCapAllocation(self):
        if len(self.FactorRanks) >= 0:
            self.CapitalAllocation.loc[self.CurrentTime] = 0
            self.Position.loc[self.CurrentTime] = 0
            #self.PositionWOSL.loc[self.CurrentTime] = 0
            self.FactorRanks = self.FactorRanks.sort_values()
            if self.ModelName == 'MonthlyMACD':
                longStocks = self.FactorRanks.index#All +ve MACD Long
                self.CapitalAllocation.loc[self.CurrentTime, longStocks] = self.CurrentNAV/self.StocksCount  #self.CurrentNAV/len(longStocks)
                if 'Liquid' in self.Close.columns:# Allocate Remain to Liquid
                    self.CapitalAllocation.loc[self.CurrentTime, 'Liquid'] = self.CurrentNAV*(1 - len(longStocks)/self.StocksCount)
                    self.Position.loc[self.CurrentTime, 'Liquid'] = 1
            else:
                maxcnt = self.SelectionCount[1] #50
                mincnt = self.SelectionCount[0] #30
                
                laststocksholding = self.CapitalAllocation.loc[self.LastTime]
                laststocksholding = laststocksholding[laststocksholding != 0].dropna()
                
                commontolast = set.intersection(set(laststocksholding.index), set(self.FactorRanks[-maxcnt:].index))
                if len(commontolast) == mincnt:
                    longStocks = commontolast
                else:
                    newEntry = self.FactorRanks[-maxcnt:].loc[list(set.difference(set(self.FactorRanks[-maxcnt:].index), commontolast))].dropna()
                    newEntry = newEntry.sort_values()
                    newEntry = newEntry.iloc[-mincnt+len(commontolast):].index
                    longStocks = list(set.union(set(commontolast), set(newEntry)))
                    
                #longStocks = self.FactorRanks.index
                self.CapitalAllocation.loc[self.CurrentTime, longStocks] = self.CurrentNAV/mincnt#len(longStocks)
                self.Position.loc[self.CurrentTime, longStocks] = 1
                if 'Liquid' in self.Close.columns:
                    self.CapitalAllocation.loc[self.CurrentTime, 'Liquid'] = self.CurrentNAV*(1 - len(longStocks)/mincnt)#self.StocksCount
                    self.Position.loc[self.CurrentTime, 'Liquid'] = 1
            
            self.UpdateOrderBook(strategyID = self.ModelName + self.IndexName.split(' ')[0])
            print(self.CurrentTime.date(), len(longStocks), self.IndexName.split(' ')[0])
                
if __name__=='__main__':
    import pickle
    import os
    import csv
    from openpyxl import load_workbook
    
    picklePath = 'G:/Shared drives/QuantFunds/EquityPlus/DataPickles/MomentumModels'
    f = open(picklePath+ '20240911.pkl', 'rb')#datetime.datetime.today().date().strftime('%Y%m%d') +'.pkl', 'rb')
    #f = open(picklePath+ datetime.datetime.today().date().strftime('%Y%m%d') +'.pkl', 'rb')
    mydata = pickle.load(f)
    f.close()
    
    if 'BSEALL INDEX' in mydata.indexprice.columns:
        mydata.indexprice['BSEALL INDEX'] = mydata.indexprice['BSE500 INDEX']
    #modelName = 'WeeklyMACD'
    rebalType = 'q'#'var' #'w', 'm', 'q'
    
    Category = rebalType.capitalize()+'Rebal'
    basePath = os.path.join('G:/Shared drives/QuantFunds/EquityPlus/Backtests/Momentum/Testing/', Category + datetime.datetime.now().date().strftime('_%d%b%Y/'))
    if not os.path.exists(basePath):
        os.makedirs(basePath)
    
    indices = {1 : 'NIFTY', 2: 'BSE100', 3: 'BSE200', 4: 'BSE500', 5: 'SPBSAIP',  6: 'BSEALL'}
    numStocks = {'NIFTY': [15, 20], 'BSE100': [20, 30], 'BSE200': [30, 50], 'BSE500': [30, 50], 'SPBSAIP': [20, 30], 'BSEALL' : [30, 50]}
    portFileName = basePath + 'ModelsData.xlsx'
    #ModelNum = {'3MPriceMom' : [3, 4], '1YrPriceMomVolAdj': [2, 3], '6MPriceMom' : [3, 4], '1YrPriceMom' : [3, 4], 'DrawDownAvgLow': [3], '1YrPriceLowVol' : [3, 4], '6MPriceLowVol': [2], 'MonthlyMACD': [4], 'Near52Wkhigh_NrTop15Pct': [4]}
    ModelNum = {'6MPriceMom' : [3, 4, 6]}#{'1YrPriceMomVolAdj': [2, 3], '6MPriceLowVol': [2], '6MPriceMom' : [3, 4, 6]}
    #ModelNum = {'1YrPriceMomVolAdj': [2, 3, 5], '6MPriceMom' : [3, 4, 5], '1YrPriceMom' : [3, 4], '6MPriceLowVol': [2, 3]}
    #ModelNum = {'3MPriceMom' : 2, '1YrPriceMomVolAdj': 1, '6MPriceMom' : 2, 'DrawDownAvgLow': 3, '1YrPriceLowVol' : 4, '6MPriceLowVol': 5, 'MonthlyMACD': 6, 'Near52Wkhigh_NrTop15Pct': 7}
    
    #indices = {1: ['BSE100', 'BSE200'], 2: ['BSE200', 'BSE500'], 3: ['BSE200'], 4: ['BSE200', 'BSE500'], 5: ['BSE100'], 6: ['BSE500'], 7: ['BSE500']}
    #indices = ['NIFTY INDEX', 'BSE100 INDEX', 'BSE200 INDEX', 'BSE500 INDEX', 'NSEMCAP INDEX', 'NSESMCP INDEX']
    for modelName, num in ModelNum.items():
        print(modelName, ' :Running!')
        indexList = [indices[id] for id in num] #indices[num]
        for index in indexList:
            indexName = index+ ' INDEX'
            mydata.indexcomponents = mydata.IndexCompDict[indexName]
            #mydata.Close.Cash = (mydata.Close.Cash + 0.065/252).cumprod()
            if 'TTMT/A IN' in mydata.Close.columns:
                del mydata.Close['TTMT/A IN']
                
            model = PriceMomentum(mydata, indexName, modelName, rebalType = rebalType, stocksCount = numStocks[index])
            model.run()
            backtestName = Category+'_'+modelName+'_'+indexName.split(' ')[0]+'_Top'+"_".join([str(it) for it in numStocks[index]])+'_'#'_Top30_'
            model.ResultFrameWithIndex()
            filepath = os.path.join(basePath, backtestName+'.xlsx')
            model.savebacktestresult(filepath, fullData = False)
            
            navData = model.PlotResult.loc[:, ['NAV', indexName]].dropna()
            yrsDiff = (navData.index[-1] - navData.index[0]).days/365.0
            cagrRet = (navData.iloc[-1])**(1/yrsDiff) -1
            AverageChurn = int(100*model.Churn.resample('a').sum().mean())
            titlename = 'CAGR-'+ str(["%.1f" % i for i in (cagrRet.values*100)]) + ',Churn-'+str(AverageChurn)+'%, ' +backtestName
            
            navData.plot(title = titlename, figsize =(18,6))
            plt.savefig(os.path.join(basePath, backtestName+'_NAV.jpg'))
            
            temp = navData.pct_change(252)
            temp = 100*(temp[temp.columns[0]] - temp[temp.columns[1]]).dropna()
            temp = pandas.DataFrame(temp)
            temp.columns = ['Rolling 1Yr Returns ' + titlename]
            temp['X-Axis'] = 0
            temp.plot(title = titlename, figsize =(18,6))
            plt.savefig(os.path.join(basePath, backtestName+'_RR.jpg'))
            
            tmp = navData.resample('a').last()
            tmp = tmp.pct_change()
            tmp.index = tmp.index.year
            tmp.plot(kind = 'bar', title = titlename, figsize =(18,6))
            plt.savefig(os.path.join(basePath, backtestName+'_YPlot.jpg'))
            
            tmp2 = navData.resample('q').last()
            tmp2 = tmp2.pct_change()
            tmp2.index = [i.strftime('%b-%y') for i in tmp2.index]
            tmp2.plot(kind = 'bar', title = titlename, figsize =(18,6))
            plt.savefig(os.path.join(basePath, backtestName+'_QPlot.jpg'))
            
            tradeDF = model.trade_reg.get_trade_register()
                
            filepath = os.path.join(basePath, backtestName+'_TradesRegister.xlsx')
            stats_obj = Stats(tradeDF)
            statsSymbolDF = stats_obj.create_stats(filter_by = Filter.SYMBOL)
            statsPositionDF = stats_obj.create_stats(filter_by = Filter.POSITION)
            
            writer = pandas.ExcelWriter(filepath)
            tradeDF.to_excel(writer,'Trades')
            statsSymbolDF.transpose().to_excel(writer,'Stats-Symbol')
            statsPositionDF.to_excel(writer,'Stats-Position')
            writer.save()
            writer.close()
            
            monthlyReturns = model.NAV.resample('m').last().pct_change().loc[datetime.date(2023, 1, 1):model.LastTime, :]
            monthlyReturns.columns = ['MonthlyRets']
            monthlyReturns.index = monthlyReturns.index.strftime('%b%Y')
            
            overAllPosition = model.Position.resample('m' if model.RebalType == 'var' else model.RebalType).last().loc[:model.LastTime, :]            
            LastMonth = overAllPosition.iloc[-2]
            LastMonth = LastMonth[LastMonth>0]
            
            CurMonth = overAllPosition.iloc[-1]
            CurMonth = CurMonth[CurMonth>0]
            
            changes = overAllPosition.diff(1).iloc[-1]
            changes = changes[changes !=0]
            
            exits = changes[changes<0]
            entry = changes[changes>0]
            
            try:
                book = load_workbook(portFileName)
                fwriter = pandas.ExcelWriter(portFileName, engine = 'openpyxl')
                fwriter.book = book
            except:
                fwriter = pandas.ExcelWriter(portFileName, engine = 'openpyxl')
            
            
            #worksheet = workbook.create_sheet(modelName +'_'+index)            
            pandas.DataFrame(LastMonth.index, columns = ['LastMonth']).to_excel(fwriter, modelName +'_'+index, startrow = 0, startcol = 0, index = False)#LastMonth.name.date().strftime('%d%b%Y')
            pandas.DataFrame(CurMonth.index, columns = ['CurrentMonth']).to_excel(fwriter, modelName +'_'+index, startrow = 0, startcol = 3, index = False)#CurMonth.name.date().strftime('%d%b%Y')
            pandas.DataFrame(exits.index, columns = ['Exits']).to_excel(fwriter, modelName +'_'+index, startrow = 0, startcol = 6, index = False)
            pandas.DataFrame(entry.index, columns = ['Entry']).to_excel(fwriter, modelName +'_'+index, startrow = 0, startcol = 8, index = False)
            model.FactorRanks.to_excel(fwriter, modelName +'Ranks', index =True)
            monthlyReturns.to_excel(fwriter, modelName +'_'+index, startrow = 0, startcol = 12, index = True)            
            fwriter.save()
            fwriter.close()           
            
            AnnualVol = navData.pct_change().std()*numpy.sqrt(252)
            retStats = [Category, modelName, indexName, navData.index[0].date(), navData.index[-1].date(), cagrRet.loc['NAV'], cagrRet.loc[indexName], AnnualVol.loc['NAV'], AnnualVol.loc[indexName], AverageChurn/100]
            
            statsFileName = basePath +'StatsFile.csv' #'G:/Shared drives/QuantFunds/EquityPlus/Backtests/Momentum/StatsFile.csv'
            fields = ['Category', 'ModelName', 'IndexName', 'StartDate', 'EndDate', 'ModelCAGR','IndexCAGR',  'ModelVol', 'IndexVol', 'Churn']
            if not os.path.exists(statsFileName):
                with open(statsFileName, 'w', newline='') as fp:
                    writer = csv.writer(fp)
                    writer.writerow(fields)
            with open(statsFileName, 'a', newline='') as fp:
                writer = csv.writer(fp)
                writer.writerow(retStats)